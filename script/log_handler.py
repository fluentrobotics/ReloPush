import os
from openpyxl import Workbook

# Get the directory of the current script
current_directory = os.path.dirname(__file__)

# Function to create Excel file for each list of dictionaries
def create_excel_file(data):
    for sublist in data:
        data_name = sublist[0]['data_name'] # just use one in the first data
        mode_name = sublist[0]['mode']
        
        wb = Workbook()
        ws = wb.active
        keys = sublist[0].keys()
        
        # Write headers
        for col, key in enumerate(keys, start=1):
            ws.cell(row=1, column=col, value=key)
        
        # Write data
        for row, entry in enumerate(sublist, start=2):
            for col, key in enumerate(keys, start=1):
                ws.cell(row=row, column=col, value=entry[key])
        
        # Save Excel file with data_name as file name
        current_directory = os.path.dirname(__file__)
        file_name = (data_name+'_'+mode_name)
        file_path = os.path.join(os.path.dirname(current_directory) + "/log", f'{file_name}.xlsx')
        print(file_path)
        wb.save(file_path)

folder = "../log/raw"


# Construct the full file path
folder_path = os.path.join(current_directory, folder)

#print(os.listdir(folder_path))



multiple_file_data = []

for filename in os.listdir(folder_path):
    file_path = os.path.join(folder_path, filename)
    if os.path.isfile(file_path):
        single_file_data = []
        with open(file_path, 'r') as file:
            file_data = file.read().strip().split('#')
            for block in file_data:
                if block.strip():
                    lines = block.strip().split('\n')
                    inner_dict = {}
                    for line in lines:
                        key, value = line.split(':', 1)
                        inner_dict[key.strip()] = value.strip()
                    single_file_data.append(inner_dict)
            multiple_file_data.append(single_file_data)

#save
output_folder = folder_path
create_excel_file(multiple_file_data)


print("data handled")